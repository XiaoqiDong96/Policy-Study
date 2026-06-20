#!/usr/bin/env python3
"""Create TXT audit files, review folders, and an Excel workbook for NEV LLM results."""

from __future__ import annotations

import argparse
import json
import re
import shutil
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def read_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not path.exists():
        return rows
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            if line.strip():
                rows.append(json.loads(line))
    return rows


def safe_name(value: Any, max_len: int = 80) -> str:
    text = str(value or "").strip()
    text = re.sub(r"[\\/:*?\"<>|\r\n\t]+", "_", text)
    text = re.sub(r"\s+", " ", text)
    return (text[:max_len] or "untitled").strip(" ._")


def as_joined(values: Any, sep: str = "；") -> str:
    if values is None:
        return ""
    if isinstance(values, list):
        return sep.join(str(v) for v in values if str(v).strip())
    if isinstance(values, dict):
        return json.dumps(values, ensure_ascii=False)
    return str(values)


def confidence_value(cls: Dict[str, Any]) -> float:
    vals = [
        cls.get("classification_confidence"),
        cls.get("confidence_is_industrial_policy"),
        cls.get("confidence_is_nev_related"),
    ]
    nums = []
    for val in vals:
        try:
            nums.append(float(val))
        except Exception:
            pass
    return min(nums) if nums else 0.0


def policy_yes_score(cls: Dict[str, Any]) -> float:
    try:
        value = float(cls.get("confidence_is_industrial_policy"))
    except Exception:
        value = 1.0 if cls.get("is_industrial_policy") else 0.0
    return max(0.0, min(1.0, value))


def boundary_uncertainty_score(cls: Dict[str, Any]) -> float:
    score = policy_yes_score(cls)
    return max(0.0, min(1.0, 1.0 - 2.0 * abs(score - 0.5)))


def is_boundary_uncertain(cls: Dict[str, Any], low: float = 0.25, high: float = 0.75) -> bool:
    score = policy_yes_score(cls)
    return low <= score <= high


def problem_reasons(item: Dict[str, Any]) -> List[str]:
    cls = item.get("classification") or {}
    consensus = item.get("adversarial_consensus") or {}
    reasons: List[str] = []

    if consensus.get("classification_disagreement"):
        reasons.append("模型投票不一致")
    if int(consensus.get("models_failed") or 0) > 0:
        reasons.append("存在模型调用失败")
    vote_share = consensus.get("policy_vote_share")
    try:
        vote_share_f = float(vote_share)
        if vote_share_f not in (0.0, 1.0):
            reasons.append("政策判断非全票一致")
    except Exception:
        pass
    if is_boundary_uncertain(cls):
        reasons.append(
            f"产业政策yes-score处于边界区间({policy_yes_score(cls):.2f}); "
            f"boundary_uncertainty={boundary_uncertainty_score(cls):.2f}"
        )
    risk = str(cls.get("false_positive_risk") or "").lower()
    if risk in {"medium", "high", "中", "高"}:
        reasons.append(f"误判风险={cls.get('false_positive_risk')}")
    if cls.get("is_industrial_policy") and not cls.get("policy_tools"):
        if cls.get("measure_specificity") != "guidance_only":
            reasons.append("判为产业政策但政策工具为空")
    if item.get("llm_error"):
        reasons.append("LLM错误字段非空")

    return reasons


def is_non_unanimous_vote(item: Dict[str, Any]) -> bool:
    consensus = item.get("adversarial_consensus") or {}
    try:
        vote_share = float(consensus.get("policy_vote_share"))
    except Exception:
        return False
    return 0.0 < vote_share < 1.0


def txt_body(item: Dict[str, Any], idx: int, classification_available: bool) -> str:
    cls = item.get("classification") or {}
    consensus = item.get("adversarial_consensus") or {}
    lines = [
        f"序号: {idx}",
        f"id: {item.get('id', '')}",
        f"title: {item.get('title', '')}",
        f"province: {item.get('province', '')}",
        f"pub_depart: {item.get('pub_depart', '')}",
        f"law_type: {item.get('law_type', '')}",
        f"pub_num: {item.get('pub_num', '')}",
        f"pub_date: {item.get('pub_date', '')}",
        f"use_date: {item.get('use_date', '')}",
        f"date_month: {item.get('date_month', '')}",
        f"detail_url: {item.get('detail_url', '')}",
        f"candidate_score: {item.get('candidate_score', '')}",
        f"nev_keyword_hits: {as_joined(item.get('nev_keyword_hits'))}",
        f"policy_keyword_hits: {as_joined(item.get('policy_keyword_hits'))}",
    ]
    if classification_available:
        lines.extend(
            [
                "",
                "=== LLM 分类结果 ===",
                f"is_nev_related: {cls.get('is_nev_related', '')}",
                f"is_industrial_policy: {cls.get('is_industrial_policy', '')}",
                f"classification_confidence: {cls.get('classification_confidence', '')}",
                f"confidence_is_nev_related: {cls.get('confidence_is_nev_related', '')}",
                f"confidence_is_industrial_policy: {cls.get('confidence_is_industrial_policy', '')}",
                f"false_positive_risk: {cls.get('false_positive_risk', '')}",
                f"policy_tone: {cls.get('policy_tone', '')}",
                f"timing: {cls.get('timing', '')}",
                f"policy_side: {cls.get('policy_side', '')}",
                f"measure_specificity: {cls.get('measure_specificity', '')}",
                f"direct_target_evidence: {cls.get('direct_target_evidence', '')}",
                f"measure_or_guidance_evidence: {cls.get('measure_or_guidance_evidence', '')}",
                f"policy_tools: {as_joined(cls.get('policy_tools'))}",
                f"tool_groups: {as_joined(cls.get('tool_groups'))}",
                f"target_segments: {as_joined(cls.get('target_segments'))}",
                f"specific_measures: {as_joined(cls.get('specific_measures'))}",
                f"strength_score: {cls.get('strength_score', '')}",
                f"coverage_breadth_score: {cls.get('coverage_breadth_score', '')}",
                f"policy_yes_votes: {consensus.get('policy_yes_votes', '')}",
                f"policy_vote_share: {consensus.get('policy_vote_share', '')}",
                f"classification_disagreement: {consensus.get('classification_disagreement', '')}",
                f"adversarial_not_policy_case: {cls.get('adversarial_not_policy_case', '')}",
                f"decision_reason: {cls.get('decision_reason', '')}",
                f"llm_error: {item.get('llm_error', '')}",
            ]
        )
    lines.extend(["", "=== LLM 输入正文片段 ===", str(item.get("llm_body") or "")])
    return "\n".join(lines).rstrip() + "\n"


def result_row(
    item: Dict[str, Any],
    idx: int,
    txt_path: Path,
    review_path: Path | None,
    non_unanimous_path: Path | None,
) -> Dict[str, Any]:
    cls = item.get("classification") or {}
    consensus = item.get("adversarial_consensus") or {}
    reasons = problem_reasons(item)
    return {
        "seq": idx,
        "id": item.get("id"),
        "title": item.get("title"),
        "province": item.get("province"),
        "pub_depart": item.get("pub_depart"),
        "law_type": item.get("law_type"),
        "pub_date": item.get("pub_date"),
        "date_month": item.get("date_month"),
        "candidate_score": item.get("candidate_score"),
        "nev_keyword_hits": as_joined(item.get("nev_keyword_hits")),
        "policy_keyword_hits": as_joined(item.get("policy_keyword_hits")),
        "is_nev_related": cls.get("is_nev_related"),
        "is_industrial_policy": cls.get("is_industrial_policy"),
        "classification_confidence": cls.get("classification_confidence"),
        "confidence_is_industrial_policy": cls.get("confidence_is_industrial_policy"),
        "policy_yes_score": policy_yes_score(cls),
        "boundary_uncertainty_score": boundary_uncertainty_score(cls),
        "false_positive_risk": cls.get("false_positive_risk"),
        "policy_tone": cls.get("policy_tone"),
        "timing": cls.get("timing"),
        "policy_side": cls.get("policy_side"),
        "measure_specificity": cls.get("measure_specificity"),
        "direct_target_evidence": cls.get("direct_target_evidence"),
        "measure_or_guidance_evidence": cls.get("measure_or_guidance_evidence"),
        "policy_tools": as_joined(cls.get("policy_tools")),
        "tool_groups": as_joined(cls.get("tool_groups")),
        "target_segments": as_joined(cls.get("target_segments")),
        "specific_measures": as_joined(cls.get("specific_measures")),
        "strength_score": cls.get("strength_score"),
        "coverage_breadth_score": cls.get("coverage_breadth_score"),
        "policy_yes_votes": consensus.get("policy_yes_votes"),
        "policy_vote_share": consensus.get("policy_vote_share"),
        "classification_disagreement": consensus.get("classification_disagreement"),
        "models_succeeded": consensus.get("models_succeeded"),
        "models_failed": consensus.get("models_failed"),
        "problem_flag": bool(reasons),
        "problem_reasons": as_joined(reasons),
        "non_unanimous_vote_flag": is_non_unanimous_vote(item),
        "decision_reason": cls.get("decision_reason"),
        "adversarial_not_policy_case": cls.get("adversarial_not_policy_case"),
        "llm_error": item.get("llm_error"),
        "detail_url": item.get("detail_url"),
        "txt_path": str(txt_path),
        "review_txt_path": str(review_path or ""),
        "non_unanimous_txt_path": str(non_unanimous_path or ""),
    }


def model_rows(item: Dict[str, Any], idx: int) -> Iterable[Dict[str, Any]]:
    for model_result in item.get("model_classifications") or []:
        cls = model_result.get("classification") or {}
        yield {
            "seq": idx,
            "id": item.get("id"),
            "title": item.get("title"),
            "model": model_result.get("model"),
            "company": model_result.get("company"),
            "error": model_result.get("error"),
            "is_nev_related": cls.get("is_nev_related"),
            "is_industrial_policy": cls.get("is_industrial_policy"),
            "classification_confidence": cls.get("classification_confidence"),
            "confidence_is_industrial_policy": cls.get("confidence_is_industrial_policy"),
            "policy_yes_score": policy_yes_score(cls),
            "boundary_uncertainty_score": boundary_uncertainty_score(cls),
            "false_positive_risk": cls.get("false_positive_risk"),
            "policy_tone": cls.get("policy_tone"),
            "timing": cls.get("timing"),
            "policy_side": cls.get("policy_side"),
            "measure_specificity": cls.get("measure_specificity"),
            "direct_target_evidence": cls.get("direct_target_evidence"),
            "measure_or_guidance_evidence": cls.get("measure_or_guidance_evidence"),
            "policy_tools": as_joined(cls.get("policy_tools")),
            "decision_reason": cls.get("decision_reason"),
            "adversarial_not_policy_case": cls.get("adversarial_not_policy_case"),
        }


def write_sheet(ws: Any, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        ws.append(["empty"])
        return
    headers = list(rows[0])
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, min(ws.max_row, 200) + 1):
            max_len = max(max_len, len(str(ws.cell(row=row_idx, column=col_idx).value or "")))
        width = min(max(max_len + 2, 10), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook(
    output_path: Path,
    result_rows: List[Dict[str, Any]],
    problem_rows_list: List[Dict[str, Any]],
    model_vote_rows: List[Dict[str, Any]],
) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    policy_count = sum(1 for r in result_rows if r.get("is_industrial_policy") is True)
    problem_count = sum(1 for r in result_rows if r.get("problem_flag") is True)
    non_unanimous_count = sum(1 for r in result_rows if r.get("non_unanimous_vote_flag") is True)
    tone_counts = Counter(r.get("policy_tone") for r in result_rows if r.get("is_industrial_policy") is True)
    measure_counts = Counter(
        r.get("measure_specificity") for r in result_rows if r.get("is_industrial_policy") is True
    )
    tool_counts = Counter()
    for row in result_rows:
        if row.get("is_industrial_policy") is True:
            for tool in str(row.get("policy_tools") or "").split("；"):
                if tool:
                    tool_counts[tool] += 1

    summary_rows = [
        {"metric": "candidate_rows", "value": len(result_rows)},
        {"metric": "industrial_policy_rows", "value": policy_count},
        {"metric": "problem_review_rows", "value": problem_count},
        {"metric": "non_unanimous_vote_rows", "value": non_unanimous_count},
        {
            "metric": "non_unanimous_vote_ratio",
            "value": non_unanimous_count / len(result_rows) if result_rows else 0,
        },
        {"metric": "non_problem_rows", "value": len(result_rows) - problem_count},
    ]
    for key, val in tone_counts.items():
        summary_rows.append({"metric": f"tone_{key}", "value": val})
    for key, val in measure_counts.items():
        summary_rows.append({"metric": f"measure_specificity_{key}", "value": val})
    for key, val in tool_counts.most_common(20):
        summary_rows.append({"metric": f"tool_{key}", "value": val})
    write_sheet(summary, summary_rows)

    ws = wb.create_sheet("Results")
    write_sheet(ws, result_rows)
    ws = wb.create_sheet("Problems")
    write_sheet(ws, problem_rows_list)
    ws = wb.create_sheet("ModelVotes")
    write_sheet(ws, model_vote_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # Verification: make sure workbook opens and expected sheets exist.
    loaded = load_workbook(output_path, read_only=True)
    expected = {"Summary", "Results", "Problems", "ModelVotes"}
    missing = expected - set(loaded.sheetnames)
    if missing:
        raise RuntimeError(f"Workbook verification failed, missing sheets: {missing}")
    if loaded["Results"].max_row < len(result_rows) + 1:
        raise RuntimeError("Workbook verification failed, Results row count too small")
    loaded.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--candidates", required=True)
    parser.add_argument("--classified", default="")
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--excel-name", default="nev_llm_results.xlsx")
    args = parser.parse_args()

    candidates_path = Path(args.candidates)
    classified_path = Path(args.classified) if args.classified else None
    out_dir = Path(args.output_dir)
    txt_dir = out_dir / "txt_all"
    review_dir = out_dir / "review_problem_cases"
    non_unanimous_dir = out_dir / "review_non_unanimous_vote_cases"
    txt_dir.mkdir(parents=True, exist_ok=True)
    review_dir.mkdir(parents=True, exist_ok=True)
    non_unanimous_dir.mkdir(parents=True, exist_ok=True)

    source_rows = read_jsonl(classified_path) if classified_path and classified_path.exists() else read_jsonl(candidates_path)
    classification_available = bool(classified_path and classified_path.exists() and source_rows and source_rows[0].get("classification"))

    result_rows: List[Dict[str, Any]] = []
    problem_rows_list: List[Dict[str, Any]] = []
    model_vote_rows: List[Dict[str, Any]] = []

    for idx, item in enumerate(source_rows, start=1):
        name = f"{idx:04d}_id{safe_name(item.get('id'), 24)}_{safe_name(item.get('title'), 70)}.txt"
        txt_path = txt_dir / name
        txt_path.write_text(txt_body(item, idx, classification_available), encoding="utf-8")

        reasons = problem_reasons(item) if classification_available else []
        review_path = None
        if reasons:
            review_path = review_dir / name
            shutil.copy2(txt_path, review_path)
        non_unanimous_path = None
        if classification_available and is_non_unanimous_vote(item):
            non_unanimous_path = non_unanimous_dir / name
            shutil.copy2(txt_path, non_unanimous_path)

        row = result_row(item, idx, txt_path, review_path, non_unanimous_path)
        result_rows.append(row)
        if reasons:
            problem_rows_list.append(row)
        model_vote_rows.extend(model_rows(item, idx))

    if classification_available:
        build_workbook(out_dir / args.excel_name, result_rows, problem_rows_list, model_vote_rows)
    else:
        (out_dir / "candidate_txt_manifest.json").write_text(
            json.dumps(
                {
                    "candidate_rows": len(result_rows),
                    "txt_dir": str(txt_dir),
                    "note": "Classification not available yet; Excel workbook is generated after LLM classification.",
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

    print(json.dumps(
        {
            "rows": len(result_rows),
            "txt_dir": str(txt_dir),
            "review_dir": str(review_dir),
            "non_unanimous_dir": str(non_unanimous_dir),
            "problem_rows": len(problem_rows_list),
            "non_unanimous_vote_rows": sum(1 for r in result_rows if r.get("non_unanimous_vote_flag") is True),
            "excel": str(out_dir / args.excel_name) if classification_available else "",
        },
        ensure_ascii=False,
    ))


if __name__ == "__main__":
    main()
