#!/usr/bin/env python3
"""Build strict project-level MOE enterprise-university collaboration edges.

Only tables that expose both an enterprise column and a university column are
accepted.  Duplicate enterprise-sorted/university-sorted attachments collapse
on the official project ID.  Enterprise cities remain blank unless a city is
explicit in the enterprise string.
"""

from __future__ import annotations

import json
import math
import re
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from priority_worker_utils import (
    PROJECT, ConservativeCityLocator, blank_297_grid, compact, load_cities,
    read_csv, stable_id, write_csv,
)


EVENTS = PROJECT / "01_source_register" / "download_events.csv"
OFFICIAL_LISTS = PROJECT / "05_intermediate" / "official_lists_records.csv"
EDGES = PROJECT / "05_intermediate" / "moe_collaboration_project_edges.csv"
FILE_QC = PROJECT / "10_qc" / "moe_collaboration_file_qc.csv"
PANEL = PROJECT / "06_panel" / "moe_collaboration_edges_297_city_year_2012_2026.csv"
SUMMARY = PROJECT / "10_qc" / "moe_collaboration_edges_summary.json"

COMPANY_HEADERS = ("公司名称", "企业名称", "立项企业名称", "支持公司", "支持企业")
UNIVERSITY_HEADERS = ("承担学校", "承担高校", "学校名称", "高校名称", "合作高校")
ID_HEADERS = ("项目编号", "项目编码")
NAME_HEADERS = ("项目名称", "立项项目名称")
TYPE_HEADERS = ("项目类型", "项目类别")
LEADER_HEADERS = ("项目负责人", "负责人")
AUDITED_SUFFIXES = {".doc", ".docx", ".pdf", ".xlsx"}
PROJECT_TYPE_PREFIXES = tuple(sorted({
    "新工科、新医科、新农科、新文科建设", "新工科建设", "新工科建设项目",
    "教学内容和课程体系改革", "教学内容和课程体系改革项目",
    "实践条件和实践基地建设", "实践条件和实践基地建设项目",
    "师资培训", "师资培训项目", "创新创业教育改革", "创新创业教育改革项目",
    "创新创业联合基金", "创新创业联合基金项目",
}, key=len, reverse=True))

EDGE_FIELDS = [
    "edge_id", "project_key", "project_id", "project_year", "company_name",
    "university_name", "project_type", "project_name", "project_leader",
    "company_city_code", "company_city_name", "company_city_match_method",
    "company_city_match_evidence", "university_city_code", "university_city_name",
    "university_city_match_method", "same_city_edge", "cross_city_edge",
    "source_file", "source_url", "source_sha256", "extraction_method",
    "duplicate_source_count", "formal_variable_eligible",
]

QC_FIELDS = [
    "source_file", "source_url", "file_suffix", "parser_status", "table_count",
    "qualifying_table_count", "raw_project_rows", "accepted_project_rows",
    "rejected_blank_entity_rows", "detected_project_years", "coverage_note", "error",
]

PANEL_FIELDS = [
    "city_code", "city_name", "province_code", "province_name", "year",
    "moe_collaboration_source_covered_year", "moe_collaboration_project_count",
    "moe_collaboration_unique_company_count", "moe_collaboration_company_city_mapped_count",
    "moe_collaboration_same_city_edge_count", "moe_collaboration_cross_city_edge_count",
    "moe_collaboration_unique_partner_city_count",
    "moe_collaboration_cross_city_edge_share_of_mapped",
    "moe_collaboration_partner_city_shannon",
    "moe_collaboration_partner_city_hhi",
]


def header_index(cells: list[Any], aliases: tuple[str, ...]) -> int | None:
    normalized = [compact(value) for value in cells]
    for index, value in enumerate(normalized):
        if any(alias == value or alias in value for alias in aliases):
            return index
    return None


def detect_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]] | None:
    for row_number, cells in enumerate(rows[:12]):
        company = header_index(cells, COMPANY_HEADERS)
        university = header_index(cells, UNIVERSITY_HEADERS)
        if company is None or university is None:
            continue
        mapping = {"company": company, "university": university}
        for field, aliases in (
            ("project_id", ID_HEADERS), ("project_name", NAME_HEADERS),
            ("project_type", TYPE_HEADERS), ("leader", LEADER_HEADERS),
        ):
            index = header_index(cells, aliases)
            if index is not None:
                mapping[field] = index
        return row_number, mapping
    return None


def cell(row: list[Any], index: int | None) -> str:
    return compact(row[index]) if index is not None and index < len(row) else ""


def iter_docx_tables(path: Path) -> Iterable[tuple[str, list[list[Any]]]]:
    from docx import Document
    document = Document(path)
    for number, table in enumerate(document.tables, start=1):
        yield f"python-docx:table_{number}", [[item.text for item in row.cells] for row in table.rows]


def iter_xlsx_tables(path: Path) -> Iterable[tuple[str, list[list[Any]]]]:
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            yield f"openpyxl:{worksheet.title}", rows
    finally:
        workbook.close()


def pdf_header_field(value: str) -> str:
    value = compact(value)
    for field, aliases in (
        ("project_id", ID_HEADERS), ("company", COMPANY_HEADERS),
        ("university", UNIVERSITY_HEADERS), ("project_type", TYPE_HEADERS),
        ("project_name", NAME_HEADERS), ("leader", LEADER_HEADERS),
    ):
        if any(alias == value or alias in value for alias in aliases):
            return field
    return ""


def pdf_columns(words: list[dict[str, Any]]) -> tuple[float, dict[str, float]] | None:
    header_words = [
        (pdf_header_field(word["text"]), float(word["top"]), (float(word["x0"]) + float(word["x1"])) / 2)
        for word in words
        if pdf_header_field(word["text"])
    ]
    for field, top, _ in header_words:
        if field != "company":
            continue
        same_line = [(other_field, center) for other_field, other_top, center in header_words if abs(other_top - top) <= 3]
        mapping = {other_field: center for other_field, center in same_line}
        if "company" in mapping and "university" in mapping and "project_id" in mapping:
            return top, mapping
    return None


def iter_pdf_tables(path: Path) -> Iterable[tuple[str, list[list[Any]]]]:
    """Recover text-layer PDF rows from Poppler's preserved table layout."""
    process = subprocess.run(
        ["pdftotext", "-layout", str(path), "-"],
        check=True, capture_output=True, text=True, encoding="utf-8",
        errors="replace", timeout=600,
    )
    last_columns: dict[str, float] | None = None
    for page_number, page_text in enumerate(process.stdout.split("\f"), start=1):
        lines = page_text.splitlines()
        header_index = -1
        columns: dict[str, float] | None = None
        for line_number, line in enumerate(lines):
            positions: dict[str, float] = {}
            for field, aliases in (
                ("project_id", ID_HEADERS), ("company", COMPANY_HEADERS),
                ("university", UNIVERSITY_HEADERS), ("project_type", TYPE_HEADERS),
                ("project_name", NAME_HEADERS), ("leader", LEADER_HEADERS),
            ):
                for alias in aliases:
                    start = line.find(alias)
                    if start >= 0:
                        positions[field] = start + len(alias) / 2
                        break
            if {"project_id", "company", "university"}.issubset(positions):
                header_index, columns = line_number, positions
                last_columns = positions
                break
        if columns is None:
            if last_columns is None:
                continue
            columns = last_columns
            header_index = -1
        anchors: list[tuple[str, int, tuple[int, int]]] = []
        for line_number, line in enumerate(lines):
            if line_number <= header_index:
                continue
            match = re.search(r"(?<!\d)(20\d{10,16})(?!\d)", line)
            if match:
                anchors.append((match.group(1), line_number, match.span(1)))
        if not anchors:
            continue
        ordered_columns = sorted(columns.items(), key=lambda item: item[1])
        boundaries: list[int] = []
        for index in range(len(ordered_columns) - 1):
            boundary = int(round((ordered_columns[index][1] + ordered_columns[index + 1][1]) / 2))
            if ordered_columns[index + 1][0] == "leader":
                # Poppler's monospace layout places short leader names several
                # character cells left of the centered header label.
                boundary = min(boundary, int(round(ordered_columns[index + 1][1] - 8)))
            boundaries.append(boundary)
        assigned: dict[str, dict[str, list[tuple[int, str]]]] = {
            project_id: defaultdict(list) for project_id, _, _ in anchors
        }
        anchor_spans = {line_number: span for _, line_number, span in anchors}

        def split_fields(original: str, span: tuple[int, int] | None = None) -> dict[str, str]:
            line = original
            if span:
                start, end = span
                line = line[:start] + " " * (end - start) + line[end:]
            starts = [0] + boundaries
            ends = boundaries + [max(len(line), boundaries[-1] + 1)]
            return {
                field: compact(line[start:end])
                for (field, _), start, end in zip(ordered_columns, starts, ends)
                if field != "project_id" and compact(line[start:end])
            }

        base_fields = {
            project_id: split_fields(lines[line_number], span)
            for project_id, line_number, span in anchors
        }
        for line_number, original in enumerate(lines):
            if line_number <= header_index:
                continue
            distances = [
                (identifier, abs(line_number - anchor_line))
                for identifier, anchor_line, _ in anchors
            ]
            distance = min(item[1] for item in distances)
            if distance > 2:
                continue
            values = split_fields(original, anchor_spans.get(line_number))
            nearest = [identifier for identifier, candidate_distance in distances if candidate_distance == distance]
            if len(nearest) > 1 and values:
                project_id = max(
                    nearest,
                    key=lambda identifier: sum(
                        bool(value)
                        and len(base_fields[identifier].get(field, "")) <= 1
                        for field, value in values.items()
                    ),
                )
            else:
                project_id = nearest[0]
            for field, value in values.items():
                if value:
                    assigned[project_id][field].append((line_number, value))
        canonical = ["项目编号", "公司名称", "承担学校", "项目类型", "项目名称", "项目负责人"]
        rows: list[list[str]] = [canonical]
        for project_id, _, _ in anchors:
            values = assigned[project_id]
            rows.append([
                project_id,
                "".join(value for _, value in sorted(values.get("company", []))),
                "".join(value for _, value in sorted(values.get("university", []))),
                "".join(value for _, value in sorted(values.get("project_type", []))),
                "".join(value for _, value in sorted(values.get("project_name", []))),
                "".join(value for _, value in sorted(values.get("leader", []))),
            ])
        yield f"pdftotext_layout_table:page_{page_number}", rows


def iter_doc_audit(path: Path) -> Iterable[tuple[str, list[list[Any]]]]:
    """Audit legacy DOC text; yield only if a strict pipe table is present."""
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from build_full_collection_text_index import extract_text

    text, method = extract_text(path, "application/msword")
    if not (
        re.search(r"项目编号", text)
        and re.search(r"公司名称|企业名称|支持公司", text)
        and re.search(r"承担学校|承担高校|学校名称", text)
    ):
        return
    # No known legacy file currently reaches this branch.  Refuse to infer
    # columns from prose; a future true table must be handled explicitly.
    yield f"{method}:strict_headers_detected_but_parser_unavailable", []


def university_map() -> dict[str, dict[str, str]]:
    candidates: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in read_csv(OFFICIAL_LISTS):
        if (
            row.get("source_id") == "moe_universities"
            and row.get("entity_name") and row.get("prefecture_code")
            and row.get("usable_for_city_panel") == "1"
        ):
            candidates[compact(row["entity_name"])].append(row)
    selected: dict[str, dict[str, str]] = {}
    for name, rows in candidates.items():
        if len({row["prefecture_code"] for row in rows}) == 1:
            selected[name] = max(rows, key=lambda row: int(row.get("list_year") or 0))
    return selected


def infer_year(project_id: str, context: str) -> int:
    match = re.match(r"(20\d{2})", project_id)
    if match and 2012 <= int(match.group(1)) <= 2026:
        return int(match.group(1))
    years = [int(value) for value in re.findall(r"20\d{2}", context) if 2012 <= int(value) <= 2026]
    return min(years) if years else 0


def repair_pdf_project_type(project_type: str, project_name: str) -> tuple[str, str]:
    combined = compact(project_type) + compact(project_name)
    for prefix in PROJECT_TYPE_PREFIXES:
        position = combined.find(prefix)
        if 0 <= position <= 20:
            return prefix, combined[:position] + combined[position + len(prefix):]
    return project_type, project_name


def reconcile_project_duplicates(
    duplicates: list[dict[str, Any]],
    universities: dict[str, dict[str, str]],
) -> dict[str, Any]:
    """Reconcile paired enterprise-sorted and university-sorted official rows."""
    def most_common(values: list[str]) -> str:
        return Counter(values).most_common(1)[0][0] if values else ""

    base = max(duplicates, key=lambda item: (
        compact(item["university_name"]) in universities,
        bool(re.search(r"(?:公司|集团|研究院|中心|学会)$", compact(item["company_name"]))),
        item["project_type"] in PROJECT_TYPE_PREFIXES,
        len(compact(item["company_name"])),
    ))
    row = dict(base)

    company_values = list({compact(item["company_name"]) for item in duplicates if compact(item["company_name"])})
    standard_companies = [
        value for value in company_values
        if re.search(r"(?:公司|集团|研究院|中心|学会)$", value)
    ]
    row["company_name"] = max(standard_companies or company_values, key=len)

    university_values = [compact(item["university_name"]) for item in duplicates if compact(item["university_name"])]
    exact_universities = [value for value in university_values if value in universities]
    plausible_universities = [
        value for value in university_values
        if re.search(r"(?:大学|学院|学校)(?:[（(].*[）)])?$", value)
    ]
    if exact_universities:
        row["university_name"] = most_common(exact_universities)
    else:
        embedded_official = {
            official_name
            for value in university_values
            for official_name in universities
            if len(official_name) >= 4 and official_name in value
        }
        if embedded_official:
            row["university_name"] = max(embedded_official, key=len)
        elif plausible_universities:
            row["university_name"] = min(set(plausible_universities), key=len)

    def clean_name_spill(item: dict[str, Any]) -> str:
        name = compact(item["project_name"])
        candidate_company = compact(item["company_name"])
        selected_company = compact(row["company_name"])
        if not name or len(candidate_company) >= len(selected_company):
            return name
        prefix = 0
        while prefix < len(candidate_company) and candidate_company[prefix] == selected_company[prefix]:
            prefix += 1
        suffix = 0
        while (
            suffix < len(candidate_company) - prefix
            and candidate_company[-1 - suffix] == selected_company[-1 - suffix]
        ):
            suffix += 1
        candidate_without_gap = selected_company[:prefix] + (selected_company[len(selected_company) - suffix:] if suffix else "")
        missing = selected_company[prefix:len(selected_company) - suffix if suffix else len(selected_company)]
        if candidate_without_gap == candidate_company and missing and name.startswith(missing):
            return name[len(missing):]
        return name

    exact_types = [item["project_type"] for item in duplicates if item["project_type"] in PROJECT_TYPE_PREFIXES]
    if exact_types:
        row["project_type"] = most_common(exact_types)
    project_names: list[str] = []
    for item in duplicates:
        cleaned_name = clean_name_spill(item)
        if cleaned_name:
            project_names.append(cleaned_name)
    if project_names:
        row["project_name"] = max(project_names, key=len)
    leaders = [compact(item["project_leader"]) for item in duplicates if compact(item["project_leader"])]
    plausible_leaders = [value for value in leaders if len(value) <= 30]
    if plausible_leaders:
        row["project_leader"] = min(plausible_leaders, key=len)
    if len(duplicates) > 1:
        row["extraction_method"] = f"{row['extraction_method']}|field_reconciled_by_official_project_id"
    return row


def main() -> int:
    cities = load_cities()
    city_by_code = {row["city_code"]: row for row in cities}
    locator = ConservativeCityLocator(cities)
    universities = university_map()

    registered: dict[str, dict[str, str]] = {}
    for event in read_csv(EVENTS):
        suffix = Path(event.get("saved_path", "")).suffix.lower()
        if (
            event.get("source_id") == "moe_industry_university_projects"
            and event.get("http_status", "").startswith("2")
            and event.get("saved_path") and suffix in AUDITED_SUFFIXES
        ):
            registered[event["saved_path"]] = event

    candidates: list[dict[str, Any]] = []
    qc_rows: list[dict[str, Any]] = []
    for relative, event in sorted(registered.items()):
        path = PROJECT / relative
        qc: dict[str, Any] = {
            "source_file": relative, "source_url": event.get("url", ""),
            "file_suffix": path.suffix.lower(), "table_count": 0,
            "qualifying_table_count": 0, "raw_project_rows": 0,
            "accepted_project_rows": 0, "rejected_blank_entity_rows": 0,
        }
        if not path.is_file():
            qc.update({"parser_status": "failed", "error": "registered_file_missing"})
            qc_rows.append(qc)
            continue
        try:
            if path.suffix.lower() == ".docx":
                iterator = iter_docx_tables(path)
            elif path.suffix.lower() == ".xlsx":
                iterator = iter_xlsx_tables(path)
            elif path.suffix.lower() == ".pdf":
                iterator = iter_pdf_tables(path)
            else:
                iterator = iter_doc_audit(path)
            for method, rows in iterator:
                qc["table_count"] += 1
                detected = detect_header(rows)
                if not detected:
                    continue
                qc["qualifying_table_count"] += 1
                header_row, columns = detected
                context = " ".join(compact(value) for row in rows[:header_row + 1] for value in row)
                for row in rows[header_row + 1:]:
                    company = cell(row, columns["company"])
                    university = cell(row, columns["university"])
                    if not any(compact(value) for value in row):
                        continue
                    qc["raw_project_rows"] += 1
                    if not company or not university or company in COMPANY_HEADERS or university in UNIVERSITY_HEADERS:
                        qc["rejected_blank_entity_rows"] += 1
                        continue
                    project_id = cell(row, columns.get("project_id"))
                    project_name = cell(row, columns.get("project_name"))
                    project_type = cell(row, columns.get("project_type"))
                    if method.startswith("pdftotext_layout_table"):
                        project_type, project_name = repair_pdf_project_type(project_type, project_name)
                    leader = cell(row, columns.get("leader"))
                    year = infer_year(project_id, f"{context} {relative} {event.get('url', '')}")
                    if not year:
                        qc["rejected_blank_entity_rows"] += 1
                        continue
                    key = project_id or stable_id(year, company, university, project_name, project_type)
                    candidates.append({
                        "project_key": key, "project_id": project_id, "project_year": year,
                        "company_name": company, "university_name": university,
                        "project_type": project_type, "project_name": project_name,
                        "project_leader": leader, "source_file": relative,
                        "source_url": event.get("url", ""), "source_sha256": event.get("sha256", ""),
                        "extraction_method": method,
                    })
                    qc["accepted_project_rows"] += 1
            file_years = sorted({
                int(row["project_year"]) for row in candidates
                if row["source_file"] == relative
            })
            qc["detected_project_years"] = "|".join(map(str, file_years))
            qc["parser_status"] = "parsed" if qc["qualifying_table_count"] else "not_a_company_university_project_table"
            qc["coverage_note"] = (
                "strict_company_and_university_columns_parsed"
                if qc["qualifying_table_count"]
                else "audited_attachment_has_no_strict_project_id_company_university_table"
            )
        except Exception as exc:
            qc.update({"parser_status": "failed", "error": f"{type(exc).__name__}:{exc}"})
        qc_rows.append(qc)

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in candidates:
        grouped[str(row["project_key"])].append(row)

    edges: list[dict[str, Any]] = []
    explicit_university_cache: dict[str, list[tuple[dict[str, str], str, str]]] = {}
    company_location_cache: dict[str, list[tuple[dict[str, str], str, str]]] = {}
    for key, duplicates in sorted(grouped.items()):
        row = reconcile_project_duplicates(duplicates, universities)
        university_name = compact(row["university_name"])
        mapped = universities.get(university_name)
        university_method = "exact_official_moe_university_crosswalk" if mapped else ""
        if not mapped:
            if university_name not in explicit_university_cache:
                explicit_university_cache[university_name] = locator.locate(university_name)
            explicit = explicit_university_cache[university_name]
            if len(explicit) == 1:
                city, method, _ = explicit[0]
                mapped = {"prefecture_code": city["city_code"]}
                university_method = f"{method}_in_university_name"
        university_code = mapped.get("prefecture_code", "") if mapped else ""
        university_city = city_by_code.get(university_code, {})

        if row["company_name"] not in company_location_cache:
            company_location_cache[row["company_name"]] = locator.locate(row["company_name"])
        company_matches = company_location_cache[row["company_name"]]
        company_city: dict[str, str] = {}
        company_method = ""
        company_evidence = ""
        if len(company_matches) == 1:
            company_city, company_method, company_evidence = company_matches[0]
        same_city = int(bool(company_city and university_code and company_city["city_code"] == university_code))
        cross_city = int(bool(company_city and university_code and company_city["city_code"] != university_code))
        edges.append({
            **row,
            "edge_id": stable_id("moe_collaboration", key),
            "company_city_code": company_city.get("city_code", ""),
            "company_city_name": company_city.get("city_name", ""),
            "company_city_match_method": company_method,
            "company_city_match_evidence": company_evidence,
            "university_city_code": university_code,
            "university_city_name": university_city.get("city_name", ""),
            "university_city_match_method": university_method,
            "same_city_edge": same_city, "cross_city_edge": cross_city,
            "duplicate_source_count": len(duplicates),
            "formal_variable_eligible": int(bool(university_code and 2012 <= int(row["project_year"]) <= 2026)),
        })

    covered_years = {int(row["project_year"]) for row in edges if row.get("university_city_code")}
    aggregate: dict[tuple[str, int], dict[str, Any]] = defaultdict(dict)
    bucket_edges: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in edges:
        if row.get("university_city_code"):
            bucket_edges[(row["university_city_code"], int(row["project_year"]))].append(row)
    for key, rows in bucket_edges.items():
        partner_counts = Counter(
            row["company_city_code"] for row in rows if row["company_city_code"]
        )
        mapped_count = sum(partner_counts.values())
        partner_probabilities = [
            count / mapped_count for count in partner_counts.values()
        ] if mapped_count else []
        cross_city_count = sum(int(row["cross_city_edge"]) for row in rows)
        aggregate[key] = {
            "moe_collaboration_project_count": len(rows),
            "moe_collaboration_unique_company_count": len({row["company_name"] for row in rows}),
            "moe_collaboration_company_city_mapped_count": mapped_count,
            "moe_collaboration_same_city_edge_count": sum(int(row["same_city_edge"]) for row in rows),
            "moe_collaboration_cross_city_edge_count": cross_city_count,
            "moe_collaboration_unique_partner_city_count": len(partner_counts),
            "moe_collaboration_cross_city_edge_share_of_mapped": round(cross_city_count / mapped_count, 8) if mapped_count else "",
            "moe_collaboration_partner_city_shannon": round(-sum(value * math.log(value) for value in partner_probabilities), 8) if partner_probabilities else "",
            "moe_collaboration_partner_city_hhi": round(sum(value * value for value in partner_probabilities), 8) if partner_probabilities else "",
        }
    value_fields = PANEL_FIELDS[6:]
    panel_rows = blank_297_grid(
        cities, covered_years, aggregate, value_fields,
        "moe_collaboration_source_covered_year",
    )
    for panel_row in panel_rows:
        if not int(panel_row["moe_collaboration_company_city_mapped_count"] or 0):
            panel_row["moe_collaboration_cross_city_edge_share_of_mapped"] = ""
            panel_row["moe_collaboration_partner_city_shannon"] = ""
            panel_row["moe_collaboration_partner_city_hhi"] = ""

    write_csv(EDGES, edges, EDGE_FIELDS)
    write_csv(FILE_QC, qc_rows, QC_FIELDS)
    write_csv(PANEL, panel_rows, PANEL_FIELDS)
    summary = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "status": "PASS" if edges and len(panel_rows) == 4455 else "FAIL",
        "registered_candidate_attachments": len(registered),
        "attachment_audit_by_suffix": {
            suffix: sum(Path(path).suffix.lower() == suffix for path in registered)
            for suffix in sorted(AUDITED_SUFFIXES)
        },
        "qualifying_files": sum(row.get("qualifying_table_count", 0) > 0 for row in qc_rows),
        "raw_accepted_rows_before_deduplication": len(candidates),
        "unique_project_edges": len(edges),
        "duplicate_rows_collapsed": len(candidates) - len(edges),
        "projects_with_university_city": sum(bool(row["university_city_code"]) for row in edges),
        "projects_with_explicit_company_city": sum(bool(row["company_city_code"]) for row in edges),
        "projects_retained_with_unmatched_company_city": sum(not bool(row["company_city_code"]) for row in edges),
        "cross_city_edges": sum(int(row["cross_city_edge"]) for row in edges),
        "covered_years": sorted(covered_years),
        "year_gap_evidence": {
            str(year): "covered_by_strict_project_edges" if year in covered_years else "no_strict_project_edge_rows_after_all_candidate_attachments_audited"
            for year in range(2015, 2022)
        },
        "panel_rows": len(panel_rows),
        "hard_gate": {
            "only_tables_with_company_and_university_columns": True,
            "official_project_id_deduplication": True,
            "unmatched_company_city_retained": True,
            "no_external_company_headquarters_inference": True,
        },
    }
    SUMMARY.parent.mkdir(parents=True, exist_ok=True)
    SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False))
    return 0 if summary["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
